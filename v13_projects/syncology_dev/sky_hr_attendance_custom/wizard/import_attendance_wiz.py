from odoo import models, fields, api, _
from odoo.exceptions import UserError
from openpyxl import load_workbook
import base64
from io import BytesIO
from odoo.exceptions import UserError
from datetime import datetime
import pytz


class AttendanceImportWizard(models.TransientModel):
    _name = "import.attendance.wiz"
    _description = "Import Attendance Wiz"

    xlsx_file = fields.Binary('File', help='Upload file in xlsx format')
    xlsx_file_name = fields.Char('File Name')

    def import_attendance(self):
        """
        This method will be for import Attendance
        --------------------------------------------------------------------------
        @param self: object pointer
        """
        if not (self.xlsx_file_name.endswith('.xlsx')
                or self.xlsx_file_name.endswith('.xls')):
            raise UserError(_("Please insert a valid file"))

        wb = load_workbook(filename=BytesIO(base64.b64decode(self.xlsx_file)),
                           read_only=True)
        ws = wb.active
        emp_obj = self.env['hr.employee']
        att_dict = {}
        local = pytz.timezone(self.env.context.get('tz'))
        for record in ws.iter_rows(min_row=2,
                                   max_row=None,
                                   min_col=None,
                                   max_col=None,
                                   values_only=True):
            state_index = 3 if len(record) > 3 else 2
            if state_index >= len(record) or record[state_index] not in [
                    'I', 'O'
            ]:
                continue
            if record[0]:
                employee = emp_obj.search([('id', '=', int(record[0]))])
                if not employee.ids and len(record) > 3:
                    employee = emp_obj.search([('display_name', '=', record[1])
                                               ])
                if record[state_index] == 'O':
                    lcl_co = local.localize(record[state_index - 1],
                                            is_dst=None)
                    utc_co = lcl_co.astimezone(pytz.utc)
                    utc_co1 = utc_co.strftime('%Y-%m-%d %H:%M:%S')
                    attendance_record = att_dict.get(employee.id)
                    if attendance_record:
                        attendance_record.update({'check_out': utc_co1})
                    else:
                        att_dict.update({
                            employee.id: {
                                'date': record[state_index - 1].date(),
                                'check_out': utc_co1
                            }
                        })

                if record[state_index] == 'I':
                    lcl_ci = local.localize(record[state_index - 1],
                                            is_dst=None)
                    utc_ci = lcl_ci.astimezone(pytz.utc)
                    utc_ci1 = utc_ci.strftime('%Y-%m-%d %H:%M:%S')
                    attendance_record = att_dict.get(employee.id)
                    if attendance_record:
                        attendance_record.update({'check_in': utc_ci1})
                    else:
                        att_dict.update({
                            employee.id: {
                                'date': record[state_index - 1].date(),
                                'check_in': utc_ci1
                            }
                        })

        for dict_data in att_dict:
            if not dict_data:
                continue
            att_rec = att_dict.get(dict_data)
            record_date = att_rec.get('date')
            st_time = record_date.strftime('%Y-%m-%d') + ' 00:00:00'
            en_time = record_date.strftime('%Y-%m-%d') + ' 23:59:59'
            st_dt = datetime.strptime(st_time, '%Y-%m-%d %H:%M:%S')
            en_dt = datetime.strptime(en_time, '%Y-%m-%d %H:%M:%S')
            local_st_dt = local.localize(st_dt, is_dst=None)
            utc_st_dt = local_st_dt.astimezone(pytz.utc)
            local_en_dt = local.localize(en_dt, is_dst=None)
            utc_en_dt = local_en_dt.astimezone(pytz.utc)
            
            attendance = self.env['hr.attendance'].search(
                [('employee_id', '=', dict_data),
                 ('check_in', '>=', utc_st_dt), ('check_in', '<=', utc_en_dt)],
                limit=1)
            if attendance.ids:
                if att_rec.get('check_in') and att_rec.get('check_out'):
                    attendance.write({
                        'check_in': att_rec.get('check_in'),
                        'check_out': att_rec.get('check_out')
                    })
                elif att_rec.get('check_in'):
                    attendance.write({'check_in': att_rec.get('check_in')})
                elif att_rec.get('check_out'):
                    attendance.write({'check_out': att_rec.get('check_out')})

            else:
                try:
                    if att_rec.get('check_in') and att_rec.get('check_out'):
                        self.env['hr.attendance'].create({
                        'employee_id':
                        dict_data,
                        'check_in':
                        att_rec.get('check_in'),
                        'check_out':
                        att_rec.get('check_out')
                    })
                    elif att_rec.get('check_in'):
                        self.env['hr.attendance'].create({
                        'employee_id':
                        dict_data,
                        'check_in':
                        att_rec.get('check_in')
                    })
                    elif att_rec.get('check_out'):
                        self.env['hr.attendance'].create({
                        'employee_id':
                        dict_data,
                        'check_in':
                        att_rec.get('check_out'),
                        'check_out':
                        att_rec.get('check_out')
                    })
                except:
                    raise UserError(_(dict_data))

